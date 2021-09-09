from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from string import ascii_uppercase
from arrow import utcnow
import bd

# Exportación de reportes a Excel
class reporteEXCEL(object):
    """Exportar una lista de tuplas a una tabla en un archivo xlsx (EXCEL)."""
    def __init__(self, titulo, curso, cabecera, registros, nombreEXCEL, porcentaje):
        super(reporteEXCEL, self).__init__()

        self.titulo = titulo
        self.curso = curso
        self.cabecera = cabecera
        self.registros = registros
        self.nombreEXCEL = nombreEXCEL
        self.porcentaje = porcentaje
        #TOTAL DE ALUMNOS
        self.total = bd.num_alum_matriculados(self.curso)

    def Exportar(self):
        # Workbook es el contenedor para todas las demás partes del documento.
        libroTrabajo = Workbook()
        
        hoja = libroTrabajo.active
        hoja.title = (self.titulo)
        hoja.sheet_properties.tabColor = "1072BA"
        
        # Ver líneas de cuadrícula
        hoja.sheet_view.showGridLines = False

        celdaFinal = ascii_uppercase[len(self.cabecera)]
        rangoTitulo = "B2:{}3".format(celdaFinal)
        rangoCabecera = "B10:{}10".format(celdaFinal)

        centrarTexto = Alignment(horizontal="center", vertical="center")

      # ========================== TÍTULO ==========================

        hoja.merge_cells(rangoTitulo)
        celdaTitulo = hoja.cell(row=2, column=2)
        celdaTitulo.value = self.titulo.upper()
        celdaTitulo.alignment = centrarTexto
        celdaTitulo.font = Font(color="FF000000", size=11, bold=True)

      # ===================== INFORMACIÓN EXTRA ====================

        fontInformacionExtra = Font(color="707070", size=11, bold=False)

        celdaOrigen = hoja.cell(row=5, column=2)
        celdaOrigen.value = "Curso: "+self.curso
        celdaOrigen.font = fontInformacionExtra

        celdaFechaDescarga = hoja.cell(row=6, column=2)
        celdaFechaDescarga.value = "Fecha: {}".format(utcnow().to("local").format("DD/MM/YYYY"))
        celdaFechaDescarga.font = fontInformacionExtra

        celdaCantidadDescarga = hoja.cell(row=7, column=2)
        celdaCantidadDescarga.value = "Total de alumnos: {}".format(self.total)
        celdaCantidadDescarga.font = fontInformacionExtra
        
        celdaCantidadDescarga = hoja.cell(row=8, column=2)
        celdaCantidadDescarga.value = "Total de alumnos seleccionados: {}".format(len(self.registros))
        celdaCantidadDescarga.font = fontInformacionExtra

      # ================== BORDES - COLOR (CELDAS) =================

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        colorCelda = PatternFill("solid", fgColor="B0CF71")

      # ================== BORDES - COLOR (TÍTULO) =================

        filasTitulo = hoja[rangoTitulo]

        celdaInicial = filasTitulo[0][0].row
        for fila in filasTitulo:
            filaIzquierda = fila[0]
            filaDerecha = fila[-1]
            filaIzquierda.border = filaIzquierda.border + Border(left=border.left)
            filaDerecha.border = filaDerecha.border + Border(right=border.right)
            
            for celda in fila:
                if celda.row == celdaInicial:
                    celda.border = celda.border + Border(top=border.top)
                else:
                    celda.border = celda.border + Border(bottom=border.bottom)

                celda.fill = colorCelda

      # ========= DATOS - BORDES - COLOR (CABECERA - TABLA) ========

        for indice, dato in enumerate(self.cabecera, start=2):
            hoja.cell(row=10, column=indice).value = dato
            hoja.cell(row=10, column=indice).border = border
            hoja.cell(row=10, column=indice).alignment = centrarTexto
            hoja.cell(row=10, column=indice).font = Font(color="FF000000", size=10, bold=True)
            
        filasEncabezado = hoja[rangoCabecera]
        for fila in filasEncabezado:
            for celda in fila:
                celda.fill = colorCelda
                
      # ====== REGISTROS - BORDES - COLOR (REGISTROS - TABLA) ======
        
        for filaIndice, registros in enumerate(self.registros, start=11):
            q=filaIndice-11
            for columnaIndice, registro in enumerate(registros, start=2):
                hoja.cell(row=filaIndice, column=columnaIndice).value = registro
                hoja.cell(row=filaIndice, column=columnaIndice).border = border
                hoja.cell(row=filaIndice, column=columnaIndice).alignment = Alignment(horizontal="center", vertical="center")
                if q>=0:
                    if self.porcentaje[q]<=30:
                        hoja.cell(row=filaIndice, column=columnaIndice).font = Font(color="FF0000", size=10, bold=False)
                else:
                    hoja.cell(row=filaIndice, column=columnaIndice).font = Font(color="FF000000", size=12, bold=False)
            q=q+1

      # ============== AJUSTAR ANCHO (CELDAS - TABLA) ==============

        for col in hoja.columns:
            columna = [(columna.column, columna.value) for columna in col
                       if not columna.value is None]
            if columna:
                longitudMaxima = 0
                for celda in columna:
                    if len(str(celda[1])) > longitudMaxima:
                        longitudMaxima = len(celda[1])

                ajustarAncho = (longitudMaxima+1) * 1.2
                #hoja.column_dimensions[columna[0][0]].width = ajustarAncho

        try:
            # Guardar el libro actual bajo el nombre de archivo dado
            ruta="./excel/"
            libroTrabajo.save("{}{}.xlsx".format(ruta,self.nombreEXCEL))

         # +----------------------------------------+
            retornar = "Reporte generado con éxito."
         # +----------------------------------------+
        except PermissionError:
         # +------------------------------------------------+  
            retornar = "Error inesperado: Permiso denegado."
         # +------------------------------------------------+
        except:
         # +-------------------------------+  
            retornar = "Error desconocido."
         # +-------------------------------+
        finally:
            # Cerrar el libro de trabajo (Workbook)
            libroTrabajo.close()
            return retornar
