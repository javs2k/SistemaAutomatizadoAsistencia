import wx
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from string import ascii_uppercase
from arrow import utcnow
import bd
import utils
import datetime

def reporte_alumno_excel(curso, cabecera, registros, cod_alum, nombre):
    titulo = "ASISTENCIA POR SEMANA"
    nombreEXCEL = "reporte por alumno"
    # Workbook es el contenedor para todas las demás partes del documento.
    libroTrabajo = Workbook()
    
    hoja = libroTrabajo.active
    hoja.title = (titulo)
    hoja.sheet_properties.tabColor = "1072BA"

    # Ver líneas de cuadrícula
    hoja.sheet_view.showGridLines = False

    celdaFinal = ascii_uppercase[len(cabecera)]
    rangoTitulo = "B2:{}3".format(celdaFinal)
    rangoCabecera = "B10:{}10".format(celdaFinal)

    centrarTexto = Alignment(horizontal="center", vertical="center")

    # ========================== TÍTULO ==========================

    hoja.merge_cells(rangoTitulo)
    celdaTitulo = hoja.cell(row=2, column=2)
    celdaTitulo.value = titulo.upper()
    celdaTitulo.alignment = centrarTexto
    celdaTitulo.font = Font(color="FF000000", size=11, bold=True)

    # ===================== INFORMACIÓN EXTRA ====================

    fontInformacionExtra = Font(color="707070", size=11, bold=False)

    celdaOrigen = hoja.cell(row=5, column=2)
    celdaOrigen.value = "Curso: "+curso
    celdaOrigen.font = fontInformacionExtra

    celdaFechaDescarga = hoja.cell(row=6, column=2)
    celdaFechaDescarga.value = "Fecha: {}".format(utcnow().to("local").format("DD/MM/YYYY"))
    celdaFechaDescarga.font = fontInformacionExtra

    celdaCantidadDescarga = hoja.cell(row=7, column=2)
    celdaCantidadDescarga.value = "Nombre: {}".format(nombre)
    celdaCantidadDescarga.font = fontInformacionExtra
    
    celdaCantidadDescarga = hoja.cell(row=8, column=2)
    celdaCantidadDescarga.value = "Código: {}".format(cod_alum)
    celdaCantidadDescarga.font = fontInformacionExtra

    # ================== BORDES - COLOR (CELDAS) =================

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    colorCelda = PatternFill("solid", fgColor="B0CF71")

    # ================== BORDES - COLOR (TÍTULO) =================

    filasTitulo = hoja[rangoTitulo]

    celdaInicial = filasTitulo[0][0].row
    for fila in filasTitulo:
        filaIzquierda = fila[0]
        filaDerecha = fila[-1]
        filaIzquierda.border = filaIzquierda.border + Border(left=border.left)
        filaDerecha.border = filaDerecha.border + Border(right=border.right)
        
        for celda in fila:
            if celda.row == celdaInicial:
                celda.border = celda.border + Border(top=border.top)
            else:
                celda.border = celda.border + Border(bottom=border.bottom)

            celda.fill = colorCelda

    # ========= DATOS - BORDES - COLOR (CABECERA - TABLA) ========

    for indice, dato in enumerate(cabecera, start=2):
        hoja.cell(row=10, column=indice).value = dato
        hoja.cell(row=10, column=indice).border = border
        hoja.cell(row=10, column=indice).alignment = centrarTexto
        hoja.cell(row=10, column=indice).font = Font(color="FF000000", size=10, bold=True)
        
    filasEncabezado = hoja[rangoCabecera]
    for fila in filasEncabezado:
        for celda in fila:
            celda.fill = colorCelda
            
    # ====== REGISTROS - BORDES - COLOR (REGISTROS - TABLA) ======
    
    for filaIndice, registros in enumerate(registros, start=11):

        for columnaIndice, registro in enumerate(registros, start=2):
            hoja.cell(row=filaIndice, column=columnaIndice).value = registro
            hoja.cell(row=filaIndice, column=columnaIndice).border = border
            hoja.cell(row=filaIndice, column=columnaIndice).alignment = Alignment(horizontal="center",
                                                                                vertical="center")
            hoja.cell(row=filaIndice, column=columnaIndice).font = Font(color="FF000000",
                                                                        size=10, bold=False)


    # ============== AJUSTAR ANCHO (CELDAS - TABLA) ==============

    for col in hoja.columns:
        columna = [(columna.column, columna.value) for columna in col
                    if not columna.value is None]
        if columna:
            longitudMaxima = 0
            for celda in columna:
                if len(str(celda[1])) > longitudMaxima:
                    longitudMaxima = len(celda[1])

            ajustarAncho = (longitudMaxima+1) * 1.2
            #hoja.column_dimensions[columna[0][0]].width = ajustarAncho

    try:
        # Guardar el libro actual bajo el nombre de archivo dado
        ruta="./excel/"
        libroTrabajo.save("{}{}.xlsx".format(ruta,nombreEXCEL))

        # +----------------------------------------+
        retornar = "Reporte generado con éxito."
        # +----------------------------------------+
    except PermissionError:
        # +------------------------------------------------+  
        retornar = "Error inesperado: Permiso denegado."
        # +------------------------------------------------+
    except:
        # +-------------------------------+  
        retornar = "Error desconocido."
        # +-------------------------------+
    finally:
        # Cerrar el libro de trabajo (Workbook)
        libroTrabajo.close()
        
        return retornar              

class ven_reporte_alumno(wx.Frame):
    def __init__(self, parent, title, clase, cod_alum):
        wx.Frame.__init__(self, parent, id=wx.ID_ANY, title=title, pos=wx.DefaultPosition,
                            size=wx.Size(650, 700), style=wx.DEFAULT_FRAME_STYLE | wx.TAB_TRAVERSAL)

        self.SetSizeHints(wx.DefaultSize, wx.DefaultSize)
        self.SetFont(wx.Font(wx.NORMAL_FONT.GetPointSize(), 70, 90, 90, False, wx.EmptyString))
        self.SetForegroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))
        self.SetBackgroundColour(wx.SystemSettings.GetColour(wx.SYS_COLOUR_WINDOW))

        bSizer4 = wx.BoxSizer(wx.VERTICAL)

        self.m_panel6 = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,
                                    wx.TAB_TRAVERSAL)
        bSizer5 = wx.BoxSizer(wx.VERTICAL)

        self.m_staticText4 = wx.StaticText(self.m_panel6, wx.ID_ANY, u"DATOS DEL ALUMNO",
                                            wx.DefaultPosition, wx.DefaultSize, 0)
        self.m_staticText4.Wrap(-1)
        
        bSizer5.Add(self.m_staticText4, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 5)

        self.m_panel6.SetSizer(bSizer5)
        self.m_panel6.Layout()
        bSizer5.Fit(self.m_panel6)
        bSizer4.Add(self.m_panel6, 0, wx.EXPAND | wx.ALL, 5)

        self.m_panel7 = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,wx.TAB_TRAVERSAL)
        bSizer6 = wx.BoxSizer(wx.VERTICAL)

        self.m_panel7.SetSizer(bSizer6)
        self.m_panel7.Layout()
        bSizer6.Fit(self.m_panel7)
        bSizer4.Add(self.m_panel7, 0, wx.EXPAND | wx.ALL, 5)

        self.m_panel8 = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL)
        bSizer7 = wx.BoxSizer(wx.HORIZONTAL)
        
    
        n_clases_tot = bd.num_clases_semanal(clase)*16
        semana = utils.semana_actual() - 33
        self.m_staticText7 = wx.StaticText(self.m_panel8, wx.ID_ANY, u"Clase: " + clase+
                                            "\nNumero de clases totales: "+str(n_clases_tot)+
                                            "\nSemana actual: "+str(semana),
                                            wx.DefaultPosition, wx.DefaultSize, 0)
        
        
        self.m_staticText7.Wrap(-1)
        self.m_panel8.SetSizer(bSizer7)
        self.m_panel8.Layout()
        bSizer7.Fit(self.m_panel8)
        bSizer4.Add(self.m_panel8, 0, wx.ALL | wx.EXPAND, 5)
        
        #AÑADIENDO DATOS DEL ALUMNO                       

        self.panel_de_datos = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL)
        bSizer_de_alumno = wx.GridSizer( 0, 2, 0, 0 )
        
        ##DATOS DEL ALUMNO
        cod='User.'+cod_alum[:-1]+'.1'
        datos_alumno = bd.datos_alumno(cod_alum)
        nombre= datos_alumno[0]
        edad=str(datos_alumno[1])
        ciclo=str(datos_alumno[2])
        espec=datos_alumno[3]
        if espec=='i1':
            especialidad='Ing. Industrial'
        if espec=='i2':
            especialidad='Ing. de Sistemas'
                                    
        
        # self.m_staticText16 = wx.StaticText( self.m_panel6, wx.ID_ANY, u"MyLabel", wx.DefaultPosition, wx.DefaultSize, 0 )
    #	self.m_staticText16.Wrap( -1 )
    # 		bSizer4.Add( self.m_staticText16, 0, wx.ALL, 5 )
        self.panel_de_tabla = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, wx.TAB_TRAVERSAL)

        bSizer_tabla = wx.BoxSizer( wx.HORIZONTAL )
        bSizer_tabla.SetMinSize( wx.Size( -1,-1 ) )
        self.ESPACIO = wx.StaticText(self.panel_de_tabla, wx.ID_ANY,"                                                  ",
                                            wx.DefaultPosition, wx.DefaultSize, 0) 
        bSizer_tabla.Add(self.ESPACIO, 0, wx.ALL, 5)
        
        self.m_grid1 = wx.grid.Grid(self, wx.ID_ANY, wx.DefaultPosition,wx.DefaultSize , 0)

        lis = bd.consulta("SELECT fecha, hora_inicio, hora_final FROM HORARIO_CLASE WHERE cod_clase = (?)",(clase,))
        day = {'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miercoles', 'Thursday': 'Jueves', 'Friday': 'Viernes',
                'Saturday': 'Sábado', 'Sunday': 'Domingo'}
        
        col = []
        cabecera=[]
        registros_exc=[]
        for i in lis:
            col.append(day[i[0]])
            
            
        # Grid
        self.m_grid1.CreateGrid(16, len(lis)+1)
        self.m_grid1.EnableEditing(False)
        self.m_grid1.EnableGridLines(True)
        self.m_grid1.SetGridLineColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_ACTIVECAPTION ) )
        self.m_grid1.EnableDragGridSize(True)
        self.m_grid1.SetMargins(0,0)           
        
        self.m_grid1.SetColLabelValue(0, "Semana")
        cabecera.append('semana')
        q=0
        for i in range(1,len(col)+1):
            self.m_grid1.SetColLabelValue(i, col[i-1])
            #insertando dato a la calumna de excel
            cabecera.append(str(col[i-1]))
            

        for i in range(0, 16):
            aux=[]
            # añadimos la primera columna
            self.m_grid1.SetCellValue(i, 0, "Semana " + str(i+1))
            a='Semana ' + str(i+1)
            aux.append(a)
            # añadimos la segunda columna
            for j in range(0,len(col)):
                list = bd.consulta(
                    "SELECT cod_alumno FROM ASISTENCIA a, HORARIO_CLASE h WHERE a.cod_alumno=(?) and a.cod_horario_clase = h.cod_horario_clase and h.cod_clase=(?) "
                    "and DATEPART(week, a.fecha_m)=(?) and h.fecha = (?)",
                    (cod_alum, clase, i+23, lis[j][0],))
                day = {'Monday': 1, 'Tuesday': 2, 'Wednesday': 3, 'Thursday': 4, 'Friday': 5, 'Saturday': 6, 'Sunday': 7}
                
                dia_hoy = datetime.date.today().isoweekday()
                
                if i+1<=semana and dia_hoy <= day[lis[j][0]]:
                    if len(list)>=1:
                        asist='ASISTIÓ'
                        q=q+1
                    elif len(list)==0:
                        asist='FALTÓ'
                else:
                    asist='- -'

                self.m_grid1.SetCellValue(i, j+1, asist)
                aux.append(asist)
            registros_exc.append(aux)
            
        #GENERAMOS EL REPORTE EXCEL
        reporte_alumno_excel(clase,cabecera,registros_exc,cod_alum,nombre)
        
        # Columns
        self.m_grid1.AutoSizeColumns()
        self.m_grid1.EnableDragColMove(False)
        self.m_grid1.EnableDragColSize(True)
        self.m_grid1.SetColLabelSize(25)
        self.m_grid1.SetColLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Rows
        self.m_grid1.AutoSizeRows()
        self.m_grid1.EnableDragRowSize(True)
        self.m_grid1.SetRowLabelSize(80)
        self.m_grid1.SetRowLabelAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)

        # Label Appearance
        self.m_grid1.SetLabelBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_ACTIVECAPTION ) )
        self.m_grid1.SetLabelFont( wx.Font( 12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL, False, "Arial Narrow" ) )


        # Cell Defaults
        self.m_grid1.SetDefaultCellAlignment(wx.ALIGN_CENTER, wx.ALIGN_CENTER )
        bSizer_tabla.Add(self.m_grid1, 1, wx.ALL, 0)

        
        asist=q
        asist_por= round(asist*100/n_clases_tot)
        self.m_staticText8=wx.BitmapButton(self.panel_de_datos, wx.ID_ANY,wx.Bitmap('./dataset/'+cod+'.jpg', wx.BITMAP_TYPE_ANY),
                                        wx.DefaultPosition, wx.DefaultSize, wx.BU_AUTODRAW)
        bSizer_de_alumno.Add(self.m_staticText8, 0, wx.ALIGN_CENTER_HORIZONTAL | wx.ALL, 5)
        
        
        bSizer2 = wx.BoxSizer( wx.VERTICAL )                        
        
        self.m_staticText9 = wx.StaticText(self.panel_de_datos, wx.ID_ANY,u"CODIGO DE ALUMNO: "+cod_alum,
                                            wx.DefaultPosition, wx.DefaultSize, 0) 
        self.m_staticText9.Wrap(-1)
        bSizer2.Add(self.m_staticText9, 0, wx.ALL, 5)
        
        self.m_staticText10 = wx.StaticText(self.panel_de_datos, wx.ID_ANY,u"NOMBRE: "+nombre,
                                            wx.DefaultPosition, wx.DefaultSize, 0) 
        self.m_staticText10.Wrap(-1)
        bSizer2.Add(self.m_staticText10, 0, wx.ALL, 5)
        
        self.m_staticText11 = wx.StaticText(self.panel_de_datos, wx.ID_ANY,"EDAD: "+edad,
                                            wx.DefaultPosition, wx.DefaultSize, 0) 
        self.m_staticText11.Wrap(-1)
        bSizer2.Add(self.m_staticText11, 0, wx.ALL, 5)
        
        self.m_staticText12 = wx.StaticText(self.panel_de_datos, wx.ID_ANY,"CICLO: "+ciclo,
                                            wx.DefaultPosition, wx.DefaultSize, 0)
        self.m_staticText12.Wrap(-1)
        bSizer2.Add(self.m_staticText12, 0, wx.ALL, 5)
        
        self.m_staticText13 = wx.StaticText(self.panel_de_datos, wx.ID_ANY,"ESPECIALIDAD: "+especialidad+" ("+espec+")",
                                            wx.DefaultPosition, wx.DefaultSize, 0)
        self.m_staticText13.Wrap(-1)
        bSizer2.Add(self.m_staticText13, 0, wx.ALL, 5)
        
        self.m_staticText14 = wx.StaticText(self.panel_de_datos, wx.ID_ANY,"NÚMERO DE ASISTENCIAS: "+str(asist),
                                            wx.DefaultPosition, wx.DefaultSize, 0) 
        self.m_staticText14.Wrap(-1)
        bSizer2.Add(self.m_staticText14, 0, wx.ALL, 5)
        
        self.m_staticText15 = wx.StaticText(self.panel_de_datos, wx.ID_ANY,"PORCENTAJE DE ASISTENCIA TOTAL: "+str(asist_por)+"%",
                                            wx.DefaultPosition, wx.DefaultSize, 0)                                                              
        self.m_staticText15.Wrap(-1)
        bSizer2.Add(self.m_staticText15, 0, wx.ALL, 5)
        
        self.m_gauge2 = wx.Gauge( self.panel_de_datos, wx.ID_ANY, n_clases_tot, wx.DefaultPosition, wx.DefaultSize, wx.GA_HORIZONTAL )
        self.m_gauge2.SetValue( asist )
        bSizer2.Add(self.m_gauge2, 0, wx.ALL, 5 )
        
        bSizer_de_alumno.Add( bSizer2, 1, wx.EXPAND, 5 )
        self.panel_de_datos.SetSizer(bSizer_de_alumno)
        self.panel_de_datos.Layout()
        bSizer_de_alumno.Fit(self.panel_de_datos)
        bSizer4.Add(self.panel_de_datos, 0, wx.ALL | wx.EXPAND, 5)
        
        #Añadiendo al sizer
        bSizer4.Add(bSizer_tabla, 1, wx.ALL, 0)
        
        #AGREGANDO EL BOTON DE EXCEL
        self.m_panel9 = wx.Panel(self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize,wx.TAB_TRAVERSAL)
        bSizer8 = wx.BoxSizer(wx.HORIZONTAL)
        self.m_panel9.SetSizer(bSizer8)
        self.m_panel9.Layout()
        bSizer8.Fit(self.m_panel9)
        self.bexcel=wx.Button(self.m_panel9,0, u"ABRIR REPORTE EN EXCEL")                                
        self.bimprimir=wx.Button(self.m_panel9,1, u"IMPRIMIR")
        bSizer8.Add(self.bexcel, 1, wx.ALL, 10)
        bSizer8.Add(self.bimprimir, 1, wx.ALL, 10)
        bSizer4.Add(self.m_panel9, 0, wx.ALIGN_CENTER, wx.ALL, 5) 
        
        self.SetSizer(bSizer4)
        self.Layout()
        self.SetIcon(wx.Icon("./imagenes/UNI.ico"))

        self.Centre(wx.BOTH)

        self.Show()